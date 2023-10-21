# excel_processor.py
import os
import openpyxl
from utils import Colors


class ExcelProcessor:
    def __init__(self, directory: str):
        self.__directory__ = directory

    def search_files(self, field: tuple, filenames):
        column_name, values = field
        found = False
        for filename in os.listdir(self.__directory__):
            if filename.endswith(".xlsx") and not filename.startswith("~$") and filename.lower().replace('.xlsx',
                                                                                                         '') in filenames:
                try:
                    workbook = openpyxl.load_workbook(os.path.join(self.__directory__, filename))
                    for sheet in workbook.sheetnames:
                        worksheet = workbook[sheet]
                        headers = [cell.value for cell in worksheet[1]]
                        column_index = None if column_name is None else (
                            headers.index(column_name) + 1 if column_name in headers else None)
                        if column_index is not None or column_name is None:
                            for row in worksheet.iter_rows(min_row=2):
                                for cell in row:
                                    if cell.value is not None and (
                                            (column_index is None) or (cell.col_idx == column_index)):
                                        try:
                                            # 尝试将单元格值和字段都转换为整数进行比较
                                            if int(cell.value) in map(int, values):
                                                print(
                                                    Colors.GREEN + f'在文件{Colors.LIGHTYELLOW_EX}{filename:<30}{Colors.GREEN}的工作表{Colors.LIGHTYELLOW_EX}{sheet:<10}{Colors.GREEN}的第{Colors.LIGHTYELLOW_EX}{cell.row:<5}{Colors.GREEN}行第{Colors.LIGHTYELLOW_EX}{cell.column:<5}{Colors.GREEN}列查找到目标' + Colors.RESET)
                                                found = True
                                        except ValueError:
                                            # 如果转换失败，则按原来的方式进行比较
                                            if str(cell.value) in map(str, values):
                                                print(
                                                    Colors.GREEN + f'在文件{Colors.LIGHTYELLOW_EX}{filename:<30}{Colors.GREEN}的工作表{Colors.LIGHTYELLOW_EX}{sheet:<10}{Colors.GREEN}的第{Colors.LIGHTYELLOW_EX}{cell.row:<5}{Colors.GREEN}行第{Colors.LIGHTYELLOW_EX}{cell.column:<5}{Colors.GREEN}列查找到目标' + Colors.RESET)
                                                found = True
                except Exception as e:
                    print(Colors.RED + f"处理文件{filename}时发生错误: {str(e)}" + Colors.RESET)
        if not found:
            print(Colors.YELLOW + f"未在任何文件中找到字段'{field}'" + Colors.RESET)

    def update_files(self, old_field: tuple, new_field: tuple, filenames):
        old_column_name, old_values = old_field
        new_column_name, new_values = new_field
        found = False
        if (old_column_name is None and new_column_name is not None) or (
                old_column_name is not None and new_column_name is None):
            print(Colors.RED + "错误：old_column_name 和 new_column_name 必须同时为空或非空" + Colors.RESET)
            return
        for filename in os.listdir(self.__directory__):
            if filename.endswith(".xlsx") and not filename.startswith("~$") and filename.lower().replace('.xlsx',
                                                                                                         '') in filenames:
                try:
                    workbook = openpyxl.load_workbook(os.path.join(self.__directory__, filename))
                    for sheet in workbook.sheetnames:
                        worksheet = workbook[sheet]
                        headers = [cell.value for cell in worksheet[1]]

                        old_column_index = None if old_column_name is None else (
                            headers.index(old_column_name) + 1 if old_column_name in headers else None)

                        new_column_index = None if new_column_name is None else (
                            headers.index(new_column_name) + 1 if new_column_name in headers else None)

                        if old_column_index is not None or old_column_name is None:
                            for row in worksheet.iter_rows(min_row=2):
                                for cell in row:  # 修改了这里
                                    if old_column_index is not None and new_column_index is not None:
                                        # 匹配行
                                        if str(worksheet.cell(row=cell.row, column=old_column_index).value) \
                                                in map(str, old_values):
                                            # 获取新值行
                                            new_value_rows = [row for row in worksheet.iter_rows(min_row=2)
                                                              if str(worksheet.cell(row=row[0].row,
                                                                                    column=new_column_index).value)
                                                              in map(str, new_values)]
                                            if len(new_value_rows) != len(old_values):
                                                print(Colors.RED + f"错误：需更新的行数与new_values的数量不匹配: "
                                                                   f"{len(new_value_rows)}, {len(old_values)}" + Colors.RESET)
                                                print(f'new: {[i for i in old_values]}')
                                                for i in new_value_rows:
                                                    print(i)
                                                    print("\n\n")
                                                return
                                            else:
                                                try:
                                                    for old_cell, new_cell in zip(row, new_value_rows[old_values.index(str(
                                                            worksheet.cell(row=cell.row, column=old_column_index).value))]):
                                                        old_cell.value = new_cell.value
                                                    print(
                                                        Colors.GREEN + f'在文件{Colors.LIGHTYELLOW_EX}{filename:<30}{Colors.GREEN}的工作表{Colors.LIGHTYELLOW_EX}{sheet:<10}{Colors.GREEN}的第{Colors.LIGHTYELLOW_EX}{cell.row:<5}{Colors.GREEN}行更新了目标' + Colors.RESET)
                                                    found = True
                                                except ValueError:
                                                    for old_cell, new_cell in zip(row, new_value_rows[old_values.index(int(
                                                            worksheet.cell(row=cell.row, column=old_column_index).value))]):
                                                        old_cell.value = new_cell.value
                                                    print(
                                                        Colors.GREEN + f'在文件{Colors.LIGHTYELLOW_EX}{filename:<30}{Colors.GREEN}的工作表{Colors.LIGHTYELLOW_EX}{sheet:<10}{Colors.GREEN}的第{Colors.LIGHTYELLOW_EX}{cell.row:<5}{Colors.GREEN}行更新了目标' + Colors.RESET)
                                                    found = True
                                    elif old_column_index is None and new_column_index is None:
                                        # 匹配数值
                                        if cell.value is not None:  # 添加了处理空值的判断语句
                                            try:
                                                # 尝试将单元格值和字段都转换为整数进行比较
                                                if int(cell.value) in map(int, old_values):
                                                    cell.value = int(new_values[0]) if len(new_values) == 1 else int(
                                                        new_values[old_values.index(int(cell.value))])
                                                    print(
                                                        Colors.GREEN + f'在文件{Colors.LIGHTYELLOW_EX}{filename:<30}{Colors.GREEN}的工作表{Colors.LIGHTYELLOW_EX}{sheet:<10}{Colors.GREEN}的第{Colors.LIGHTYELLOW_EX}{cell.row:<5}{Colors.GREEN}行第{Colors.LIGHTYELLOW_EX}{cell.column:<5}{Colors.GREEN}列更新了目标' + Colors.RESET)
                                                    found = True
                                            except ValueError:
                                                # 如果转换失败，则按原来的方式进行比较
                                                if str(cell.value) in map(str, old_values):
                                                    cell.value = str(new_values[old_values.index(str(cell.value))])
                                                    print(
                                                        Colors.GREEN + f'在文件{Colors.LIGHTYELLOW_EX}{filename:<30}{Colors.GREEN}的工作表{Colors.LIGHTYELLOW_EX}{sheet:<10}{Colors.GREEN}的第{Colors.LIGHTYELLOW_EX}{cell.row:<5}{Colors.GREEN}行第{Colors.LIGHTYELLOW_EX}{cell.column:<5}{Colors.GREEN}列更新了目标' + Colors.RESET)
                                                    found = True
                    workbook.save(os.path.join(self.__directory__, filename))
                except Exception as e:
                    print(Colors.RED + f"处理文件{filename}时发生错误: {str(e)}" + Colors.RESET)
        if not found:
            print(Colors.YELLOW + f"未在任何文件中找到字段'{old_field}'" + Colors.RESET)

    def parse_filenames(self, filenames_str):
        filenames_str = filenames_str.upper()
        all_files = [f.lower().replace('.xlsx', '') for f in os.listdir(self.__directory__) if
                     f.endswith(".xlsx") and not f.startswith("~$")]
        if 'ALL' in filenames_str.split():
            return all_files
        elif 'EXCEPT' in filenames_str.split():
            _, except_files_str = filenames_str.split('EXCEPT')
            except_files = except_files_str.lower().split()
            result = [f for f in all_files if f not in except_files]
            return result
        else:
            result = filenames_str.lower().split()
            return result

    def copy_row(self, key_column, source_key, target_key, filenames):
        for filename in os.listdir(self.__directory__):
            if filename.endswith(".xlsx") and not filename.startswith("~$") and filename.lower().replace('.xlsx',
                                                                                                         '') in filenames:
                try:
                    workbook = openpyxl.load_workbook(os.path.join(self.__directory__, filename))
                    for sheet in workbook.sheetnames:
                        worksheet = workbook[sheet]
                        source_row = None
                        target_row = None
                        key_column_index = None
                        # 遍历标题行，找到主键列
                        for cell in worksheet[1]:
                            if str(cell.value) == str(key_column):
                                key_column_index = cell.column
                                break
                        if key_column_index is None:
                            print(Colors.YELLOW + f"在文件{filename}的表{sheet}中，没有找到列名为{key_column}的列。" + Colors.RESET)
                            continue
                        # 在主键列中查找源行和目标行
                        for row in worksheet.iter_rows(min_row=2):  # 从第二行开始遍历
                            if str(row[key_column_index - 1].value) == str(source_key):
                                source_row = row
                            if str(row[key_column_index - 1].value) == str(target_key):
                                target_row = row
                        if source_row and target_row:
                            for source_cell, target_cell in zip(source_row[1:], target_row[1:]):  # 从第二列开始复制
                                target_cell.value = source_cell.value
                            print(
                                Colors.GREEN + f"在文件{filename}的表{sheet}中，已将主键为{source_key}的行复制到主键为{target_key}的行。" + Colors.RESET)
                        elif not source_row:
                            print(Colors.YELLOW + f"在文件{filename}的表{sheet}中，没有找到主键为{source_key}的行。" + Colors.RESET)
                        elif not target_row:
                            print(Colors.YELLOW + f"在文件{filename}的表{sheet}中，没有找到主键为{target_key}的行。" + Colors.RESET)
                    workbook.save(os.path.join(self.__directory__, filename))
                except Exception as e:
                    print(Colors.RED + f"处理文件{filename}时发生错误: {str(e)}" + Colors.RESET)

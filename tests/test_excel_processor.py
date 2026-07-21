import io
import tempfile
import unittest
from contextlib import redirect_stdout
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook
from wcwidth import strip_sequences

from services.excel_processor import ExcelProcessor


class ExcelProcessorSearchTests(unittest.TestCase):
    def create_workbook(self, directory, rows):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "个人动作"
        for row in rows:
            worksheet.append(row)

        workbook_path = Path(directory, "PersonalShow.xlsx")
        workbook.save(workbook_path)
        workbook.close()

    def search(self, directory, show_detail=True):
        output = io.StringIO()
        processor = ExcelProcessor(directory)

        with (
            patch("services.detail_formatter.get_terminal_width", return_value=120),
            redirect_stdout(output),
        ):
            filename = processor.search_files(
                (None, ["春日心事"]),
                ["personalshow"],
                show_detail,
            )

        return filename, strip_sequences(output.getvalue())

    def test_groups_multiple_matches_from_the_same_row(self):
        with tempfile.TemporaryDirectory() as directory:
            self.create_workbook(directory, [
                ["Id", "Title", "Desc", "Path", "Extra", "Optional", None, None],
                [
                    7020031,
                    "春日心事",
                    "春日心事服装动作",
                    "Avatar/春日心事",
                    "春日心事",
                    None,
                    None,
                    None,
                ],
            ])

            filename, output = self.search(directory)

        self.assertEqual("PersonalShow.xlsx", filename)
        self.assertEqual(1, output.count("查找到目标"))
        self.assertEqual(1, output.count("详细信息："))
        self.assertIn("第2行（第2、3、4、5列）", output)
        self.assertNotIn("\n...\n", output)
        self.assertIn("Optional", output)
        self.assertNotIn("\n                              :", output)

    def test_keeps_matches_from_different_rows_separate(self):
        with tempfile.TemporaryDirectory() as directory:
            self.create_workbook(directory, [
                ["Id", "Title", "Desc"],
                [7020031, "春日心事", "春日心事服装动作"],
                [7020032, "春日心事", None],
            ])

            _, output = self.search(directory)

        self.assertEqual(1, output.count("查找到目标"))
        self.assertEqual(1, output.count("详细信息："))
        self.assertIn("第2行（第2、3列）", output)
        self.assertIn("第3行（第2列）", output)
        self.assertEqual(1, output.count("\n...\n"))

    def test_starts_a_new_group_for_each_worksheet(self):
        with tempfile.TemporaryDirectory() as directory:
            workbook = Workbook()
            first_worksheet = workbook.active
            first_worksheet.title = "个人动作"
            first_worksheet.append(["Id", "Title"])
            first_worksheet.append([7020031, "春日心事"])

            second_worksheet = workbook.create_sheet("表情动作")
            second_worksheet.append(["Id", "Title"])
            second_worksheet.append([7020032, "春日心事"])

            workbook_path = Path(directory, "PersonalShow.xlsx")
            workbook.save(workbook_path)
            workbook.close()

            _, output = self.search(directory, show_detail=False)

        self.assertEqual(2, output.count("查找到目标"))
        self.assertEqual(2, output.count("第2行（第2列）"))
        self.assertIn("个人动作", output)
        self.assertIn("表情动作", output)
        self.assertNotIn("\n...\n", output)


if __name__ == "__main__":
    unittest.main()

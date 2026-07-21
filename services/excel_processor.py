# excel_processor.py
import os
import openpyxl
from openpyxl.chartsheet import Chartsheet

from entities.color import Colors
from entities.command import Keyword
from services.detail_formatter import format_detail_table
from utilities.local import get_directory
from utilities.process import start_window


class ExcelProcessor:
    def __init__(self, directory: str):
        self.__directory = directory

    def search_files(self, field: tuple, filenames, show_detail: bool) -> str:
        column_name, values = field
        target_values = tuple(map(str, values))
        ret_filename = None
        found = False
        for filename in os.listdir(self.__directory):
            if filename.endswith(".xlsx") and not filename.startswith("~$") and filename.lower().replace('.xlsx',
                                                                                                         '') in filenames:
                workbook = openpyxl.load_workbook(os.path.join(self.__directory, filename))
                for sheet in workbook.sheetnames:
                    worksheet = workbook[sheet]
                    if isinstance(worksheet, Chartsheet):
                        continue
                    headers = [cell.value for cell in worksheet[1]]
                    column_index = None if column_name is None else (
                        headers.index(column_name) + 1 if column_name in headers else None)
                    if column_index is not None or column_name is None:
                        matching_rows = []
                        for row in worksheet.iter_rows(min_row=2):
                            matching_cells = [
                                cell
                                for cell in row
                                if cell.value is not None
                                and (column_index is None or cell.col_idx == column_index)
                                and any(target in str(cell.value) for target in target_values)
                            ]
                            if not matching_cells:
                                continue

                            matching_rows.append((row, matching_cells))

                        if not matching_rows:
                            continue

                        print(
                            Colors.GREEN
                            + f'在文件{Colors.LIGHTYELLOW_EX}{filename:<30}'
                              f'{Colors.GREEN}的工作表{Colors.LIGHTYELLOW_EX}{sheet:<10}'
                              f'{Colors.GREEN}中查找到目标'
                            + Colors.RESET
                        )
                        if show_detail:
                            print('详细信息：')

                        for result_index, (row, matching_cells) in enumerate(matching_rows):
                            if result_index > 0:
                                print("...")

                            matching_columns = "、".join(
                                str(cell.column)
                                for cell in matching_cells
                            )
                            print(
                                Colors.GREEN
                                + f'第{Colors.LIGHTYELLOW_EX}{matching_cells[0].row}'
                                  f'{Colors.GREEN}行（第{Colors.LIGHTYELLOW_EX}{matching_columns}'
                                  f'{Colors.GREEN}列）'
                                + Colors.RESET
                            )

                            if show_detail:
                                detail_fields = [
                                    (header, cell.value)
                                    for header, cell in zip(headers, row)
                                    if header is not None or cell.value is not None
                                ]
                                print(format_detail_table(detail_fields, target_values))

                        print("")
                        found = True
                        ret_filename = filename
        if not found:
            print(Colors.YELLOW + f"未在任何文件中找到字段'{field}'" + Colors.RESET)
        return ret_filename

    def update_files(self, old_field: tuple, new_field: tuple, filenames, require_restart: bool = False) -> str:
        def update_rows(tgt_type: type) -> bool:
            success = False
            for old_cell, new_cell in zip(row, new_value_rows[old_values.index(tgt_type(
                    worksheet.cell(row=cell.row, column=old_column_index).value))]):
                if old_cell.column == old_column_index:
                    continue
                old_cell.value = new_cell.value
                success = True
            return success

        def update_vals(tgt_type: type) -> bool:
            success = False
            if tgt_type(cell.value) in map(tgt_type, old_values):
                cell.value = tgt_type(new_values[0]) if len(new_values) == 1 else tgt_type(
                    new_values[old_values.index(tgt_type(cell.value))])
                success = True
            return success

        old_column_name, old_values = old_field
        new_column_name, new_values = new_field
        found = False
        ret_filename = None
        if (old_column_name is None and new_column_name is not None) or (
                old_column_name is not None and new_column_name is None):
            print(Colors.RED + "错误：old_column_name 和 new_column_name 必须同时为空或非空" + Colors.RESET)
            return ret_filename
        for filename in os.listdir(self.__directory):
            if filename.endswith(".xlsx") and not filename.startswith("~$") and filename.lower().replace('.xlsx',
                                                                                                         '') in filenames:
                workbook = openpyxl.load_workbook(os.path.join(self.__directory, filename))
                workbook.save(os.path.join(self.__directory, filename))

                for sheet in workbook.sheetnames:
                    worksheet = workbook[sheet]
                    headers = [cell.value for cell in worksheet[1]]

                    old_column_index = None if old_column_name is None else (
                        headers.index(old_column_name) + 1 if old_column_name in headers else None)

                    new_column_index = None if new_column_name is None else (
                        headers.index(new_column_name) + 1 if new_column_name in headers else None)

                    # 获取新值行
                    new_value_rows = [row for row in worksheet.iter_rows(min_row=2)
                                      if row[0].row and new_column_index
                                      and str(worksheet.cell(row=row[0].row,
                                                             column=new_column_index).value)
                                      in map(str, new_values)]

                    for row in worksheet.iter_rows(min_row=2):
                        updated = False
                        for cell in row:
                            if old_column_index is not None and new_column_index is not None:
                                # 匹配行
                                if str(worksheet.cell(row=cell.row, column=old_column_index).value) \
                                        in map(str, old_values):
                                    if len(new_value_rows) != len(old_values):
                                        print(Colors.RED + f"错误：需更新的行数与new_values的数量不匹配: "
                                                           f"{len(new_value_rows)}, {len(old_values)}" + Colors.RESET)
                                        return ret_filename
                                    try:
                                        if update_rows(str): updated, found = True, True
                                    except ValueError:
                                        if update_rows(int): updated, found = True, True

                            elif old_column_index is None and new_column_index is None:
                                # 匹配数值
                                if cell.value is not None:
                                    try:
                                        if update_vals(int): updated, found = True, True
                                    except ValueError:
                                        if update_vals(str): updated, found = True, True
                        if updated:
                            print(Colors.GREEN + f'在文件{Colors.LIGHTYELLOW_EX}{filename:<30}{Colors.GREEN}的工作表'
                                                 f'{Colors.LIGHTYELLOW_EX}{sheet:<10}{Colors.GREEN}的第{Colors.LIGHTYELLOW_EX}'
                                                 f'{cell.row:<5}{Colors.GREEN}行更新了目标' + Colors.RESET)
                            ret_filename = filename
                workbook.save(os.path.join(self.__directory, filename))
                if require_restart:
                    start_window(os.path.join(self.__directory, filename))
        if not found:
            print(Colors.YELLOW + f"未在任何文件中找到字段'{old_field}'" + Colors.RESET)
        return ret_filename

    def parse_filenames(self, filenames_str):

        filenames_str = filenames_str.upper()
        all_files = [f.lower().replace('.xlsx', '') for f in os.listdir(self.__directory) if
                     f.endswith(".xlsx") and not f.startswith("~$")]
        if Keyword.ALL.upper() in filenames_str.split():
            return all_files
        elif Keyword.EXCEPT.upper() in filenames_str.split():
            _, except_files_str = filenames_str.split(Keyword.EXCEPT.upper())
            except_files = except_files_str.lower().split()
            result = [f for f in all_files if f not in except_files]
            return result
        else:
            result = filenames_str.lower().split()
            return result

    def get_pure_filenames(self, filenames_str: str):
        def find_matching_filename(input_str):
            for filename in os.listdir(self.__directory):
                if filename.split(".")[0].lower() == input_str:
                    return filename
            return None

        return [(find_matching_filename(name) if find_matching_filename(name) else
                 (name + (".xlsx" if ".xlsx" not in name else "")))
                for name in filenames_str.split() if not Keyword.is_keyword(name)]

if __name__ == '__main__':
    print(ExcelProcessor(get_directory()).get_pure_filenames(input("test pure: ")))
